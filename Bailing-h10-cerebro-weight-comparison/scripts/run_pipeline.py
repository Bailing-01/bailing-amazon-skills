#!/usr/bin/env python3
"""
H10 Cerebro multi-ASIN weight comparison pipeline.
Reads multiple Cerebro xlsx files → calculates weights → outputs JSON + formatted Excel + CSV.

Usage: python run_pipeline.py '<json_params>'
  json_params: {"files": {"ASIN1": "/path/to/file1.xlsx", ...}, "rank_cutoff": 40, "formula": "exponential", "alpha": 0.15, "power_exp": 1.5}
"""

import sys
import json
import os
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from linkfox_paths import resolve_data_path, resolve_report_path, session_root

# Cerebro column indices (0-based)
COL_KEYWORD = 0
COL_SEARCH_VOL = 1
COL_SPONSORED_RANK = 20
COL_ORGANIC_RANK = 21

# Styling constants
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def safe_int(val):
    if val is None or val == '' or val == '-' or val == 0:
        return 999
    try:
        v = int(val)
        return v if v > 0 else 999
    except (ValueError, TypeError):
        try:
            v = float(val)
            return int(v) if v > 0 else 999
        except (ValueError, TypeError):
            return 999


def safe_float(val, default=0):
    if val is None or val == '' or val == '-':
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def weight_linear(sv, rank, cutoff):
    if rank >= cutoff or rank == 999:
        return 0
    return max(0, sv * (cutoff - rank))


def weight_exponential(sv, rank, cutoff, alpha):
    if rank >= cutoff or rank == 999:
        return 0
    max_mult = cutoff - 1
    return sv * max_mult * math.exp(-alpha * (rank - 1))


def weight_power(sv, rank, cutoff, power_exp):
    if rank >= cutoff or rank == 999:
        return 0
    return sv * (cutoff - rank) ** power_exp


def calc_weight(sv, rank, formula, cutoff, alpha, power_exp):
    if formula == "linear":
        return weight_linear(sv, rank, cutoff)
    elif formula == "exponential":
        return weight_exponential(sv, rank, cutoff, alpha)
    elif formula == "power":
        return weight_power(sv, rank, cutoff, power_exp)
    return 0


def read_cerebro_file(filepath, asin, formula, cutoff, alpha, power_exp):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Table"]
    keywords = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[COL_KEYWORD]:
            continue
        kw = str(row[COL_KEYWORD]).strip().lower()
        sv = safe_float(row[COL_SEARCH_VOL])
        sp_rank = safe_int(row[COL_SPONSORED_RANK])
        org_rank = safe_int(row[COL_ORGANIC_RANK])
        ad_w = calc_weight(sv, sp_rank, formula, cutoff, alpha, power_exp)
        org_w = calc_weight(sv, org_rank, formula, cutoff, alpha, power_exp)
        keywords.append({
            "keyword": kw, "search_volume": sv,
            "sponsored_rank": sp_rank if sp_rank != 999 else None,
            "organic_rank": org_rank if org_rank != 999 else None,
            "ad_weight": round(ad_w, 2),
            "organic_weight": round(org_w, 2),
            "total_weight": round(ad_w + org_w, 2),
        })
    wb.close()
    return keywords


def compute_summary(keywords):
    total_sv = sum(k["search_volume"] for k in keywords)
    total_ad = sum(k["ad_weight"] for k in keywords)
    total_org = sum(k["organic_weight"] for k in keywords)
    total_w = sum(k["total_weight"] for k in keywords)
    kw_ad = len([k for k in keywords if k["sponsored_rank"] is not None])
    kw_org = len([k for k in keywords if k["organic_rank"] is not None])
    top30 = sorted(keywords, key=lambda x: x["total_weight"], reverse=True)[:30]
    return {
        "total_keywords": len(keywords),
        "total_search_volume": round(total_sv, 2),
        "total_ad_weight": round(total_ad, 2),
        "total_organic_weight": round(total_org, 2),
        "total_weight": round(total_w, 2),
        "kw_with_ad_rank": kw_ad,
        "kw_with_org_rank": kw_org,
        "ad_weight_pct": round(total_ad / total_w * 100, 2) if total_w > 0 else 0,
        "org_weight_pct": round(total_org / total_w * 100, 2) if total_w > 0 else 0,
        "top30_keywords": top30,
    }


def build_excel(asins, asin_summaries, full_table, comparison, formula_desc, params, output_path):
    """Build formatted Excel with 4 sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws1 = wb.create_sheet("总权重对比")
    ws1["A1"] = f"H10 Cerebro 多ASIN权重对比 — 汇总"
    ws1.merge_cells("A1:G1")
    ws1["A1"].font = Font(bold=True, size=14, color="4F46E5")
    ws1["A1"].alignment = CENTER
    ws1["A2"] = f"公式: {formula_desc}  ·  rank_cutoff={params['rank_cutoff']}  ·  alpha={params.get('alpha','')}  ·  生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1.merge_cells("A2:G2")

    headers = ["ASIN", "总关键词数", "有广告排名词数", "有自然排名词数",
               "总搜索量", "总广告权重", "总自然权重", "总权重", "广告占比", "自然占比"]
    for i, h in enumerate(headers, 1):
        ws1.cell(row=4, column=i, value=h)
    for col in range(1, len(headers) + 1):
        c = ws1.cell(row=4, column=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = THIN_BORDER

    for i, asin in enumerate(asins):
        s = asin_summaries[asin]
        r = 5 + i
        ws1.cell(row=r, column=1, value=asin).font = Font(bold=True)
        ws1.cell(row=r, column=2, value=s["total_keywords"]).alignment = RIGHT
        ws1.cell(row=r, column=3, value=s["kw_with_ad_rank"]).alignment = RIGHT
        ws1.cell(row=r, column=4, value=s["kw_with_org_rank"]).alignment = RIGHT
        ws1.cell(row=r, column=5, value=s["total_search_volume"]).alignment = RIGHT
        ws1.cell(row=r, column=6, value=s["total_ad_weight"]).alignment = RIGHT
        ws1.cell(row=r, column=7, value=s["total_organic_weight"]).alignment = RIGHT
        ws1.cell(row=r, column=8, value=s["total_weight"]).alignment = RIGHT
        ws1.cell(row=r, column=9, value=f"{s['ad_weight_pct']}%").alignment = RIGHT
        ws1.cell(row=r, column=10, value=f"{s['org_weight_pct']}%").alignment = RIGHT
        for col in [5, 6, 7, 8]:
            ws1.cell(row=r, column=col).number_format = '#,##0'
        for col in range(1, len(headers) + 1):
            ws1.cell(row=r, column=col).border = THIN_BORDER
    for col_letter in "ABCDEFGHIJ":
        ws1.column_dimensions[col_letter].width = 18

    # Sheet 2: Keyword-level comparison
    ws2 = wb.create_sheet("权重对比")
    kw_headers = ["关键词", "搜索量"]
    for asin in asins:
        kw_headers.extend([f"{asin} 广告排名", f"{asin} 自然排名",
                           f"{asin} 广告权重", f"{asin} 自然权重", f"{asin} 总权重"])
    kw_headers.extend(["最大权重", "差距"])
    for i, h in enumerate(kw_headers, 1):
        ws2.cell(row=1, column=i, value=h)
    for col in range(1, len(kw_headers) + 1):
        c = ws2.cell(row=1, column=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = THIN_BORDER

    for row_idx, kw_data in enumerate(full_table, 2):
        ws2.cell(row=row_idx, column=1, value=kw_data["keyword"]).alignment = LEFT
        ws2.cell(row=row_idx, column=2, value=kw_data.get(f"{asins[0]}_sv", 0)).alignment = RIGHT
        col = 3
        weights = []
        max_w = 0
        for asin in asins:
            sp = kw_data.get(f"{asin}_sp_rank", None)
            org = kw_data.get(f"{asin}_org_rank", None)
            ad = kw_data.get(f"{asin}_ad_weight", 0)
            orgw = kw_data.get(f"{asin}_org_weight", 0)
            tw = kw_data.get(f"{asin}_total_weight", 0)
            ws2.cell(row=row_idx, column=col, value=sp if sp is not None else "-").alignment = CENTER
            ws2.cell(row=row_idx, column=col+1, value=org if org is not None else "-").alignment = CENTER
            ws2.cell(row=row_idx, column=col+2, value=ad).alignment = RIGHT
            ws2.cell(row=row_idx, column=col+3, value=orgw).alignment = RIGHT
            ws2.cell(row=row_idx, column=col+4, value=tw).alignment = RIGHT
            for c in [col+2, col+3, col+4]:
                ws2.cell(row=row_idx, column=c).number_format = '#,##0'
            weights.append(tw)
            max_w = max(max_w, tw)
            col += 5
        gap = max(weights) - min(weights) if weights else 0
        ws2.cell(row=row_idx, column=len(kw_headers)-1, value=max_w).alignment = RIGHT
        ws2.cell(row=row_idx, column=len(kw_headers), value=gap).alignment = RIGHT
        ws2.cell(row=row_idx, column=len(kw_headers)-1).number_format = '#,##0'
        ws2.cell(row=row_idx, column=len(kw_headers)).number_format = '#,##0'
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 12

    # Sheet 3: Per-ASIN Top30
    ws3 = wb.create_sheet("各ASIN Top30")
    for i, asin in enumerate(asins):
        start = 1 + i * 33
        ws3.cell(row=start, column=1, value=f"{asin} — TOP 30 关键词").font = Font(bold=True, size=12, color="4F46E5")
        ws3.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
        h3 = ["#", "关键词", "搜索量", "广告排名", "自然排名", "广告权重", "自然权重", "总权重"]
        for j, h in enumerate(h3, 1):
            ws3.cell(row=start+1, column=j, value=h)
        for col in range(1, len(h3) + 1):
            c = ws3.cell(row=start+1, column=col)
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = THIN_BORDER
        top30 = asin_summaries[asin]["top30_keywords"]
        for k, kw in enumerate(top30):
            r = start + 2 + k
            ws3.cell(row=r, column=1, value=k+1).alignment = CENTER
            ws3.cell(row=r, column=2, value=kw["keyword"]).alignment = LEFT
            ws3.cell(row=r, column=3, value=kw["search_volume"]).alignment = RIGHT
            ws3.cell(row=r, column=4, value=kw["sponsored_rank"] if kw["sponsored_rank"] else "-").alignment = CENTER
            ws3.cell(row=r, column=5, value=kw["organic_rank"] if kw["organic_rank"] else "-").alignment = CENTER
            ws3.cell(row=r, column=6, value=kw["ad_weight"]).alignment = RIGHT
            ws3.cell(row=r, column=7, value=kw["organic_weight"]).alignment = RIGHT
            ws3.cell(row=r, column=8, value=kw["total_weight"]).alignment = RIGHT
            for col in [3, 6, 7, 8]:
                ws3.cell(row=r, column=col).number_format = '#,##0'
    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 40
    for col_letter in "CDEFGH":
        ws3.column_dimensions[col_letter].width = 14

    # Sheet 4: Gap analysis
    ws4 = wb.create_sheet("关键词差距分析")
    gap_headers = ["#", "关键词", "搜索量"]
    for asin in asins:
        gap_headers.extend([f"{asin} SpR", f"{asin} OrR", f"{asin} 权重"])
    gap_headers.extend(["差距", "领先方"])
    for i, h in enumerate(gap_headers, 1):
        ws4.cell(row=1, column=i, value=h)
    for col in range(1, len(gap_headers) + 1):
        c = ws4.cell(row=1, column=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = THIN_BORDER

    for i, c in enumerate(comparison[:50]):
        r = 2 + i
        ws4.cell(row=r, column=1, value=i+1).alignment = CENTER
        ws4.cell(row=r, column=2, value=c["keyword"]).alignment = LEFT
        ws4.cell(row=r, column=3, value=c.get(f"{asins[0]}_sv", 0)).alignment = RIGHT
        ws4.cell(row=r, column=3).number_format = '#,##0'
        col = 4
        max_w = 0; leader = ""
        for asin in asins:
            sp = c.get(f"{asin}_sp_rank", None)
            org = c.get(f"{asin}_org_rank", None)
            tw = c.get(f"{asin}_total_weight", 0)
            ws4.cell(row=r, column=col, value=sp if sp is not None else "-").alignment = CENTER
            ws4.cell(row=r, column=col+1, value=org if org is not None else "-").alignment = CENTER
            ws4.cell(row=r, column=col+2, value=tw).alignment = RIGHT
            ws4.cell(row=r, column=col+2).number_format = '#,##0'
            if tw > max_w: max_w = tw; leader = asin
            col += 3
        ws4.cell(row=r, column=len(gap_headers)-1, value=c["weight_gap"]).alignment = RIGHT
        ws4.cell(row=r, column=len(gap_headers)-1).number_format = '#,##0'
        ws4.cell(row=r, column=len(gap_headers), value=leader).alignment = CENTER
    ws4.freeze_panes = "A2"
    ws4.column_dimensions["A"].width = 5
    ws4.column_dimensions["B"].width = 40
    ws4.column_dimensions["C"].width = 12

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": True, "message": "Missing JSON params argument"}))
        sys.exit(1)

    try:
        params = json.loads(sys.argv[1])
    except json.JSONDecodeError as e:
        print(json.dumps({"error": True, "message": f"Invalid JSON: {e}"}))
        sys.exit(1)

    files = params.get("files", {})
    rank_cutoff = params.get("rank_cutoff", 40)
    formula = params.get("formula", "exponential")
    alpha = params.get("alpha", 0.15)
    power_exp = params.get("power_exp", 1.5)

    if not files or len(files) < 2:
        print(json.dumps({"error": True, "message": "At least 2 ASIN files required"}))
        sys.exit(1)

    # Read and calculate
    asin_data = {}
    asin_summaries = {}
    for asin, filepath in files.items():
        if not os.path.isfile(filepath):
            print(json.dumps({"error": True, "message": f"File not found: {filepath}"}))
            sys.exit(1)
        keywords = read_cerebro_file(filepath, asin, formula, rank_cutoff, alpha, power_exp)
        asin_data[asin] = keywords
        asin_summaries[asin] = compute_summary(keywords)

    asins = list(asin_data.keys())

    # Merge keywords
    merged = defaultdict(dict)
    for asin, keywords in asin_data.items():
        for kw in keywords:
            merged[kw["keyword"]][asin] = kw

    all_asins_set = set(asins)
    common = {kw: data for kw, data in merged.items() if set(data.keys()) == all_asins_set}

    unique = {asin: 0 for asin in asins}
    for kw, data in merged.items():
        for asin in asins:
            if asin not in data:
                unique[asin] += 1

    # Comparison (common keywords by weight gap)
    comparison = []
    for kw, data in common.items():
        row = {"keyword": kw}
        weights = []
        for asin in asins:
            row[f"{asin}_sv"] = data[asin]["search_volume"]
            row[f"{asin}_ad_weight"] = data[asin]["ad_weight"]
            row[f"{asin}_org_weight"] = data[asin]["organic_weight"]
            row[f"{asin}_total_weight"] = data[asin]["total_weight"]
            row[f"{asin}_sp_rank"] = data[asin]["sponsored_rank"]
            row[f"{asin}_org_rank"] = data[asin]["organic_rank"]
            weights.append(data[asin]["total_weight"])
        row["weight_gap"] = round(max(weights) - min(weights), 2)
        comparison.append(row)
    comparison.sort(key=lambda x: x["weight_gap"], reverse=True)

    # Full table (all keywords, sorted by max weight)
    full_table = []
    for kw, data in merged.items():
        row = {"keyword": kw}
        for asin in asins:
            if asin in data:
                d = data[asin]
                row[f"{asin}_sv"] = d["search_volume"]
                row[f"{asin}_sp_rank"] = d["sponsored_rank"]
                row[f"{asin}_org_rank"] = d["organic_rank"]
                row[f"{asin}_ad_weight"] = d["ad_weight"]
                row[f"{asin}_org_weight"] = d["organic_weight"]
                row[f"{asin}_total_weight"] = d["total_weight"]
            else:
                row[f"{asin}_sv"] = 0
                row[f"{asin}_sp_rank"] = None
                row[f"{asin}_org_rank"] = None
                row[f"{asin}_ad_weight"] = 0
                row[f"{asin}_org_weight"] = 0
                row[f"{asin}_total_weight"] = 0
        full_table.append(row)
    full_table.sort(key=lambda x: max(x.get(f"{a}_total_weight", 0) for a in asins), reverse=True)

    # Formula description
    formula_desc = {
        "linear": f"W = SV * max(0, {rank_cutoff} - Rank)",
        "exponential": f"W = SV * {rank_cutoff-1} * e^(-{alpha} * (Rank-1))",
        "power": f"W = SV * ({rank_cutoff} - Rank)^{power_exp}",
    }.get(formula, "")

    # Build output JSON
    output = {
        "asins": asins,
        "params": {"rank_cutoff": rank_cutoff, "formula": formula, "alpha": alpha, "power_exp": power_exp},
        "formula_desc": formula_desc,
        "summaries": {asin: {k: v for k, v in s.items() if k != "top30_keywords"} for asin, s in asin_summaries.items()},
        "top30_per_asin": {asin: s["top30_keywords"] for asin, s in asin_summaries.items()},
        "common_keywords_count": len(common),
        "unique_keywords_count": unique,
        "comparison_top50": comparison[:50],
        "full_table_top100": full_table[:100],
    }

    import time
    ts = int(time.time() * 1000)

    # Save JSON
    json_path = resolve_data_path("weight-analysis", ts)
    os.makedirs(os.path.dirname(json_path), exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # Save formatted Excel (4 sheets)
    excel_path = resolve_data_path("weight-comparison", ts)
    try:
        build_excel(asins, asin_summaries, full_table, comparison, formula_desc, params, excel_path)
    except Exception as e:
        sys.stderr.write(f"Excel export warning: {e}\n")

    # Output per skill-output-protocol (JSON as primary)
    size = os.path.getsize(json_path)
    excel_size = os.path.getsize(excel_path) if os.path.isfile(excel_path) else 0
    print(f"JSON: {json_path} ({size} bytes)")
    print(f"Excel: {excel_path} ({excel_size} bytes)")
    print(f"Saved full response: {json_path} ({size} bytes)")


if __name__ == "__main__":
    main()
