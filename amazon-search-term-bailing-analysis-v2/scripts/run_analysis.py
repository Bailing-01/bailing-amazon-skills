#!/usr/bin/env python3
"""
亚马逊搜索词展示量份额报告 Bailing 分析 V1
主分析脚本：读取Excel → 提取层级结构 → 帕累托分析 → 输出JSON数据

用法：
  python run_analysis.py --excel <path> --step <extract|structure|pareto|all>
"""
import argparse
import json
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": True, "message": "openpyxl not installed: pip install openpyxl"}))
    sys.exit(1)


def to_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


# ====== 列索引映射（0-indexed） ======
COL_MARKET_IMP = 0    # 搜索词总展示量
COL_COUNTRY = 2       # 国家/地区
COL_SEARCH_TERM = 3   # 客户搜索词
COL_RANK = 4          # 搜索词展示量排名
COL_SHARE = 5         # 搜索词展示份额
COL_TARGET = 6        # 投放
COL_MATCH_TYPE = 7    # 匹配类型
COL_PORTFOLIO = 8     # 广告组合名称
COL_CAMPAIGN = 9      # 广告活动名称
COL_AD_GROUP = 10     # 广告组名称
COL_CURRENCY = 11     # 货币
COL_CLICKS = 12       # 点击量
COL_IMPRESSIONS = 13  # 展示量
COL_SPEND = 15        # 花费
COL_ORDERS = 17       # 7天总订单数
COL_SALES = 18        # 7天总销售额

# CVR 基准
CVR_EXCELLENT = 0.15
CVR_NORMAL = 0.12
CVR_LOW = 0.08

# 份额基准
SHARE_EXCELLENT = 0.10
SHARE_GOOD = 0.05
SHARE_VOICE = 0.01

# 品牌关键词
BRAND_KEYWORDS = [
    'thrive', 'clinique', 'maybelline', 'laura geller', 'cellmetics',
    'everlove', 'felvaris', 'judydoll', 'facelove', 'primelash',
    'prime lash', 'life girls', 'cougex', 'tubing', 'carter',
    'revlon', "l'oreal", 'loreal'
]

SHEET_NAME = '商品推广 搜索词展示量份额 报告'


def extract_data(excel_path):
    """Step 1: 读取Excel并提取层级结构数据"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return {"error": True, "message": f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"}
    ws = wb[SHEET_NAME]

    # 国家/货币追踪（每个广告组合对应的国家和货币）
    portfolio_countries = defaultdict(set)
    portfolio_currencies = defaultdict(set)

    # 按广告组合 → 搜索词聚合
    portfolio_data = defaultdict(lambda: defaultdict(lambda: {
        'spend': 0, 'clicks': 0, 'orders': 0, 'sales': 0, 'impressions': 0,
        'rank_sum': 0, 'rank_count': 0, 'share_sum': 0, 'share_count': 0,
        'market_imp_max': 0, 'market_imp_sum': 0, 'campaigns': set(), 'targets': set(), 'match_types': set(),
        'daily': defaultdict(lambda: {'impressions': 0, 'market_imp': 0, 'clicks': 0, 'orders': 0, 'spend': 0, 'sales': 0, 'rank': 0, 'rank_count': 0})
    }))

    # 同时按广告组合 → 活动 → 投放聚合
    campaign_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
        'spend': 0, 'clicks': 0, 'orders': 0, 'sales': 0, 'impressions': 0,
        'search_terms': set(), 'match_types': set(), 'targets': set()
    })))

    target_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
        'spend': 0, 'clicks': 0, 'orders': 0, 'sales': 0, 'impressions': 0,
        'search_terms': set(), 'campaigns': set(), 'match_types': set()
    })))

    row_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[COL_SEARCH_TERM] is None:
            continue
        row_count += 1

        portfolio = row[COL_PORTFOLIO] or ''
        campaign = row[COL_CAMPAIGN] or ''
        target = row[COL_TARGET] or ''
        search_term = row[COL_SEARCH_TERM] or ''
        match_type = row[COL_MATCH_TYPE] or ''

        spend = to_float(row[COL_SPEND])
        clicks = to_float(row[COL_CLICKS])
        orders = to_float(row[COL_ORDERS])
        sales = to_float(row[COL_SALES])
        impressions = to_float(row[COL_IMPRESSIONS])
        rank = to_float(row[COL_RANK])
        share = to_float(row[COL_SHARE])
        market_imp = to_float(row[COL_MARKET_IMP])
        country = row[COL_COUNTRY] or '' if len(row) > COL_COUNTRY else ''
        currency = row[COL_CURRENCY] or '' if len(row) > COL_CURRENCY else ''

        # 追踪国家/货币
        if country:
            portfolio_countries[portfolio].add(country)
        if currency:
            portfolio_currencies[portfolio].add(currency)

        # 按搜索词聚合
        sd = portfolio_data[portfolio][search_term]
        sd['spend'] += spend
        sd['clicks'] += clicks
        sd['orders'] += orders
        sd['sales'] += sales
        sd['impressions'] += impressions
        if rank > 0:
            sd['rank_sum'] += rank
            sd['rank_count'] += 1
        if share > 0:
            sd['share_sum'] += share
            sd['share_count'] += 1
        if market_imp > sd['market_imp_max']:
            sd['market_imp_max'] = market_imp
        sd['market_imp_sum'] += market_imp
        # 每日趋势数据
        date_str = str(row[1])[:10] if row[1] else 'unknown'
        sd['daily'][date_str]['impressions'] += impressions
        sd['daily'][date_str]['market_imp'] += market_imp
        sd['daily'][date_str]['clicks'] += clicks
        sd['daily'][date_str]['orders'] += orders
        sd['daily'][date_str]['spend'] += spend
        sd['daily'][date_str]['sales'] += sales
        if rank > 0:
            sd['daily'][date_str]['rank'] += rank
            sd['daily'][date_str]['rank_count'] += 1
        sd['campaigns'].add(campaign)
        sd['targets'].add(target)
        sd['match_types'].add(match_type)

        # 按活动聚合
        cd = campaign_data[portfolio][campaign][target]
        cd['spend'] += spend
        cd['clicks'] += clicks
        cd['orders'] += orders
        cd['sales'] += sales
        cd['impressions'] += impressions
        cd['search_terms'].add(search_term)
        cd['match_types'].add(match_type)

        # 按投放聚合
        td = target_data[portfolio][target][campaign]
        td['spend'] += spend
        td['clicks'] += clicks
        td['orders'] += orders
        td['sales'] += sales
        td['impressions'] += impressions
        td['search_terms'].add(search_term)
        td['campaigns'].add(campaign)
        td['match_types'].add(match_type)

    # 计算派生指标
    for portfolio, terms in portfolio_data.items():
        for term, d in terms.items():
            d['avg_rank'] = d['rank_sum'] / d['rank_count'] if d['rank_count'] > 0 else 0
            d['avg_share'] = d['impressions'] / d['market_imp_sum'] if d.get('market_imp_sum', 0) > 0 else 0
            d['cpc'] = d['spend'] / d['clicks'] if d['clicks'] > 0 else 0
            d['cvr'] = d['orders'] / d['clicks'] if d['clicks'] > 0 else 0
            d['cpa'] = d['spend'] / d['orders'] if d['orders'] > 0 else 0
            d['acos'] = d['spend'] / d['sales'] if d['sales'] > 0 else 0
            d['market_total'] = d['impressions'] / d['avg_share'] if d['avg_share'] > 0 else 0
            d['keyword'] = term
            d['portfolio'] = portfolio
            # 转set为list
            d['campaigns'] = list(d['campaigns'])
            d['targets'] = list(d['targets'])
            d['match_types'] = list(d['match_types'])

    # 序列化campaign和target数据
    campaigns_out = {}
    targets_out = {}
    for portfolio, camps in campaign_data.items():
        campaigns_out[portfolio] = {}
        targets_out[portfolio] = {}
        for camp, targs in camps.items():
            camp_total = {'spend': 0, 'clicks': 0, 'orders': 0, 'sales': 0, 'impressions': 0,
                          'search_terms': set(), 'targets': set(), 'match_types': set()}
            for targ, d in targs.items():
                camp_total['spend'] += d['spend']
                camp_total['clicks'] += d['clicks']
                camp_total['orders'] += d['orders']
                camp_total['sales'] += d['sales']
                camp_total['impressions'] += d['impressions']
                camp_total['search_terms'].update(d['search_terms'])
                camp_total['targets'].add(targ)
                camp_total['match_types'].update(d['match_types'])

                # 投放级数据
                td = target_data[portfolio][targ][camp]
                targets_out[portfolio].setdefault(targ, {
                    'spend': 0, 'clicks': 0, 'orders': 0, 'sales': 0,
                    'impressions': 0, 'search_terms': set(), 'campaigns': set(), 'match_types': set()
                })
                to = targets_out[portfolio][targ]
                to['spend'] += d['spend']
                to['clicks'] += d['clicks']
                to['orders'] += d['orders']
                to['sales'] += d['sales']
                to['impressions'] += d['impressions']
                to['search_terms'].update(d['search_terms'])
                to['campaigns'].add(camp)
                to['match_types'].update(d['match_types'])

            cvr = camp_total['orders'] / camp_total['clicks'] if camp_total['clicks'] > 0 else 0
            cpa = camp_total['spend'] / camp_total['orders'] if camp_total['orders'] > 0 else 0
            campaigns_out[portfolio][camp] = {
                'spend': round(camp_total['spend'], 2),
                'clicks': int(camp_total['clicks']),
                'orders': int(camp_total['orders']),
                'sales': round(camp_total['sales'], 2),
                'impressions': int(camp_total['impressions']),
                'search_terms_count': len(camp_total['search_terms']),
                'targets_count': len(camp_total['targets']),
                'match_types': sorted(camp_total['match_types']),
                'cvr': round(cvr, 4),
                'cpa': round(cpa, 2),
                'acos': round(camp_total['spend'] / camp_total['sales'], 4) if camp_total['sales'] > 0 else 0,
            }

        # 序列化targets
        for targ, d in targets_out[portfolio].items():
            cvr = d['orders'] / d['clicks'] if d['clicks'] > 0 else 0
            cpa = d['spend'] / d['orders'] if d['orders'] > 0 else 0
            d['spend'] = round(d['spend'], 2)
            d['clicks'] = int(d['clicks'])
            d['orders'] = int(d['orders'])
            d['sales'] = round(d['sales'], 2)
            d['impressions'] = int(d['impressions'])
            d['search_terms_count'] = len(d['search_terms'])
            d['campaigns'] = list(d['campaigns'])
            d['match_types'] = list(d['match_types'])
            d['cvr'] = round(cvr, 4)
            d['cpa'] = round(cpa, 2)
            d['acos'] = round(d['spend'] / d['sales'], 4) if d['sales'] > 0 else 0

    # 汇总
    result = {
        'row_count': row_count,
        'portfolios': {},
    }

    for portfolio, terms in portfolio_data.items():
        terms_list = list(terms.values())
        total_spend = sum(d['spend'] for d in terms_list)
        total_clicks = sum(d['clicks'] for d in terms_list)
        total_orders = sum(d['orders'] for d in terms_list)
        total_sales = sum(d['sales'] for d in terms_list)

        result['portfolios'][portfolio] = {
            'search_terms_count': len(terms_list),
            'countries': sorted(portfolio_countries.get(portfolio, set())),
            'currencies': sorted(portfolio_currencies.get(portfolio, set())),
            'total_spend': round(total_spend, 2),
            'total_clicks': int(total_clicks),
            'total_orders': int(total_orders),
            'total_sales': round(total_sales, 2),
            'cpc': round(total_spend / total_clicks, 2) if total_clicks > 0 else 0,
            'cvr': round(total_orders / total_clicks, 4) if total_clicks > 0 else 0,
            'cpa': round(total_spend / total_orders, 2) if total_orders > 0 else 0,
            'acos': round(total_spend / total_sales, 4) if total_sales > 0 else 0,
            'campaigns': campaigns_out.get(portfolio, {}),
            'targets': targets_out.get(portfolio, {}),
            'search_terms': sorted(terms_list, key=lambda x: x['orders'], reverse=True),
        }

    return result


def pareto_analysis(portfolio_data):
    """Step 3: 帕累托分析"""
    results = {}
    for portfolio, p_data in portfolio_data['portfolios'].items():
        terms = p_data['search_terms']
        n = len(terms)
        total_orders = p_data['total_orders']
        total_spend = p_data['total_spend']

        top_4_pct = max(1, int(n * 0.04))
        top_20_pct = max(1, int(n * 0.20))

        top_4 = terms[:top_4_pct]
        top_20 = terms[:top_20_pct]
        longtail = terms[top_20_pct:]

        top_4_orders = sum(d['orders'] for d in top_4)
        top_20_orders = sum(d['orders'] for d in top_20)

        # 份额分级
        share_tiers = {'>=10%': [], '5-10%': [], '1-5%': [], '0.1-1%': [], '<0.1%': [], '0%': []}
        for d in terms:
            s = d['avg_share']
            if s == 0:
                share_tiers['0%'].append(d)
            elif s < 0.001:
                share_tiers['<0.1%'].append(d)
            elif s < 0.01:
                share_tiers['0.1-1%'].append(d)
            elif s < 0.05:
                share_tiers['1-5%'].append(d)
            elif s < 0.10:
                share_tiers['5-10%'].append(d)
            else:
                share_tiers['>=10%'].append(d)

        # 长尾分类
        categories = {
            'C1_出单长尾': [],
            'C2_有点击无转化': [],
            'C3_有曝光无点击': [],
            'C4_高份额小市场': [],
            'C5_品牌竞品词': [],
            'C6_低相关泛词': [],
            'C7_零曝光僵尸词': [],
        }

        for d in longtail:
            kw = d['keyword'].lower()
            is_brand = any(bk in kw for bk in BRAND_KEYWORDS)

            if d['impressions'] == 0 and d['clicks'] == 0:
                categories['C7_零曝光僵尸词'].append(d)
            elif d['clicks'] == 0 and d['impressions'] > 0:
                categories['C3_有曝光无点击'].append(d)
            elif d['orders'] == 0 and d['clicks'] > 0:
                if is_brand:
                    categories['C5_品牌竞品词'].append(d)
                else:
                    categories['C2_有点击无转化'].append(d)
            elif d['orders'] > 0:
                if d['avg_share'] >= 0.05 and d['market_total'] < 500000:
                    categories['C4_高份额小市场'].append(d)
                elif is_brand:
                    categories['C5_品牌竞品词'].append(d)
                else:
                    categories['C1_出单长尾'].append(d)
            else:
                categories['C6_低相关泛词'].append(d)

        # 序列化份额分级
        share_dist = {}
        for tier, items in share_tiers.items():
            share_dist[tier] = {
                'count': len(items),
                'spend': round(sum(d['spend'] for d in items), 2),
                'orders': int(sum(d['orders'] for d in items)),
            }

        # 序列化长尾分类
        longtail_cats = {}
        for cat, items in categories.items():
            longtail_cats[cat] = {
                'count': len(items),
                'spend': round(sum(d['spend'] for d in items), 2),
                'clicks': int(sum(d['clicks'] for d in items)),
                'orders': int(sum(d['orders'] for d in items)),
            }

        results[portfolio] = {
            'total_keywords': n,
            'total_orders': total_orders,
            'total_spend': total_spend,
            'top_4_count': top_4_pct,
            'top_4_orders': int(top_4_orders),
            'top_4_order_pct': round(top_4_orders / total_orders, 4) if total_orders > 0 else 0,
            'top_20_count': top_20_pct,
            'top_20_orders': int(top_20_orders),
            'top_20_order_pct': round(top_20_orders / total_orders, 4) if total_orders > 0 else 0,
            'longtail_count': len(longtail),
            'share_dist': share_dist,
            'longtail_categories': longtail_cats,
            'top_4_terms': [{
                'keyword': d['keyword'],
                'orders': int(d['orders']),
                'order_pct': round(d['orders'] / total_orders, 4) if total_orders > 0 else 0,
                'cvr': round(d['cvr'], 4),
                'share': round(d['avg_share'], 6),
                'rank': round(d['avg_rank'], 1),
                'market_total': round(d['market_total'], 0),
                'spend': round(d['spend'], 2),
                'cpc': round(d['cpc'], 2),
                'cpa': round(d['cpa'], 2),
                'acos': round(d.get('acos', 0), 4),
                'clicks': int(d['clicks']),
                'sales': round(d['sales'], 2),
                'daily_trend': sorted([{
                    'date': date,
                    'market_imp': int(dd['market_imp']),
                    'impressions': int(dd['impressions']),
                    'share': round(dd['impressions'] / dd['market_imp'], 6) if dd['market_imp'] > 0 else 0,
                    'clicks': int(dd['clicks']),
                    'orders': int(dd['orders']),
                    'spend': round(dd['spend'], 2),
                    'sales': round(dd['sales'], 2),
                    'rank': round(dd['rank'] / dd['rank_count'], 1) if dd['rank_count'] > 0 else 0,
                } for date, dd in d.get('daily', {}).items()], key=lambda x: x['date']),
            } for d in top_4],
        }

    return results


def compare_periods(current_data, prev_data, period_label='当前周期', prev_label='上周期'):
    """V2: 双周期对比分析"""
    results = {}
    cur_portfolios = current_data.get('portfolios', {})
    prev_portfolios = prev_data.get('portfolios', {})

    all_portfolio_names = set(cur_portfolios.keys()) | set(prev_portfolios.keys())

    for p_name in all_portfolio_names:
        cur_p = cur_portfolios.get(p_name, {})
        prev_p = prev_portfolios.get(p_name, {})

        # 整体对比
        overall = {
            'period_label': period_label,
            'prev_label': prev_label,
            'spend': {'cur': cur_p.get('total_spend', 0), 'prev': prev_p.get('total_spend', 0),
                       'delta': round(cur_p.get('total_spend', 0) - prev_p.get('total_spend', 0), 2)},
            'orders': {'cur': cur_p.get('total_orders', 0), 'prev': prev_p.get('total_orders', 0),
                        'delta': cur_p.get('total_orders', 0) - prev_p.get('total_orders', 0)},
            'clicks': {'cur': cur_p.get('total_clicks', 0), 'prev': prev_p.get('total_clicks', 0),
                        'delta': cur_p.get('total_clicks', 0) - prev_p.get('total_clicks', 0)},
            'cvr': {'cur': cur_p.get('cvr', 0), 'prev': prev_p.get('cvr', 0),
                    'delta': round(cur_p.get('cvr', 0) - prev_p.get('cvr', 0), 4)},
            'acos': {'cur': cur_p.get('acos', 0), 'prev': prev_p.get('acos', 0),
                     'delta': round(cur_p.get('acos', 0) - prev_p.get('acos', 0), 4)},
            'cpc': {'cur': cur_p.get('cpc', 0), 'prev': prev_p.get('cpc', 0),
                    'delta': round(cur_p.get('cpc', 0) - prev_p.get('cpc', 0), 2)},
        }

        # 搜索词级别对比
        cur_terms = {}
        prev_terms = {}
        for t in cur_p.get('search_terms', []):
            cur_terms[t['keyword']] = t
        for t in prev_p.get('search_terms', []):
            prev_terms[t['keyword']] = t

        all_keywords = set(cur_terms.keys()) | set(prev_terms.keys())

        new_keywords = []
        lost_keywords = []
        compared = []
        share_up = []
        share_down = []
        rank_up = []
        rank_down = []
        acos_worse = []
        acos_better = []

        for kw in all_keywords:
            c = cur_terms.get(kw)
            p = prev_terms.get(kw)

            if c and not p:
                new_keywords.append({
                    'keyword': kw, 'spend': c.get('spend', 0), 'orders': c.get('orders', 0),
                    'share': c.get('avg_share', 0), 'rank': c.get('avg_rank', 0),
                    'cpc': c.get('cpc', 0), 'acos': c.get('acos', 0),
                })
            elif p and not c:
                lost_keywords.append({
                    'keyword': kw, 'spend': p.get('spend', 0), 'orders': p.get('orders', 0),
                    'share': p.get('avg_share', 0), 'rank': p.get('avg_rank', 0),
                })
            elif c and p:
                share_delta = c.get('avg_share', 0) - p.get('avg_share', 0)
                rank_delta = p.get('avg_rank', 0) - c.get('avg_rank', 0)  # 正=提升
                cpc_delta = c.get('cpc', 0) - p.get('cpc', 0)
                acos_delta = c.get('acos', 0) - p.get('acos', 0)
                order_delta = c.get('orders', 0) - p.get('orders', 0)
                spend_delta = c.get('spend', 0) - p.get('spend', 0)

                entry = {
                    'keyword': kw,
                    'cur_share': round(c.get('avg_share', 0), 6),
                    'prev_share': round(p.get('avg_share', 0), 6),
                    'share_delta': round(share_delta, 6),
                    'cur_rank': round(c.get('avg_rank', 0), 1),
                    'prev_rank': round(p.get('avg_rank', 0), 1),
                    'rank_delta': round(rank_delta, 1),
                    'cur_cpc': round(c.get('cpc', 0), 2),
                    'prev_cpc': round(p.get('cpc', 0), 2),
                    'cpc_delta': round(cpc_delta, 2),
                    'cur_acos': round(c.get('acos', 0), 4),
                    'prev_acos': round(p.get('acos', 0), 4),
                    'acos_delta': round(acos_delta, 4),
                    'cur_orders': int(c.get('orders', 0)),
                    'prev_orders': int(p.get('orders', 0)),
                    'order_delta': int(order_delta),
                    'cur_spend': round(c.get('spend', 0), 2),
                    'prev_spend': round(p.get('spend', 0), 2),
                    'spend_delta': round(spend_delta, 2),
                }
                compared.append(entry)

                if share_delta > 0.01:
                    share_up.append(entry)
                elif share_delta < -0.01:
                    share_down.append(entry)
                if rank_delta > 3:
                    rank_up.append(entry)
                elif rank_delta < -3:
                    rank_down.append(entry)
                if acos_delta > 0.10:
                    acos_worse.append(entry)
                elif acos_delta < -0.10:
                    acos_better.append(entry)

        # Top4核心词对比
        top4_comparison = []
        for t in cur_p.get('top_4_terms', []):
            kw = t['keyword']
            prev_t = prev_terms.get(kw)
            if prev_t:
                top4_comparison.append({
                    'keyword': kw,
                    'cur_share': round(t.get('share', 0), 6),
                    'prev_share': round(prev_t.get('avg_share', prev_t.get('share', 0)), 6),
                    'share_delta': round(t.get('share', 0) - prev_t.get('avg_share', prev_t.get('share', 0)), 6),
                    'cur_rank': round(t.get('rank', 0), 1),
                    'prev_rank': round(prev_t.get('avg_rank', prev_t.get('rank', 0)), 1),
                    'rank_delta': round(prev_t.get('avg_rank', prev_t.get('rank', 0)) - t.get('rank', 0), 1),
                    'cur_cpc': round(t.get('cpc', 0), 2),
                    'prev_cpc': round(prev_t.get('cpc', 0), 2),
                    'cur_acos': round(t.get('acos', 0), 4),
                    'prev_acos': round(prev_t.get('acos', 0), 4),
                    'cur_orders': int(t.get('orders', 0)),
                    'prev_orders': int(prev_t.get('orders', 0)),
                })

        # 排序
        share_up.sort(key=lambda x: x['share_delta'], reverse=True)
        share_down.sort(key=lambda x: x['share_delta'])
        rank_up.sort(key=lambda x: x['rank_delta'], reverse=True)
        rank_down.sort(key=lambda x: x['rank_delta'])
        acos_worse.sort(key=lambda x: x['acos_delta'], reverse=True)
        acos_better.sort(key=lambda x: x['acos_delta'])

        results[p_name] = {
            'overall': overall,
            'top4_comparison': top4_comparison,
            'new_keywords': sorted(new_keywords, key=lambda x: x.get('spend', 0), reverse=True)[:20],
            'lost_keywords': sorted(lost_keywords, key=lambda x: x.get('spend', 0), reverse=True)[:20],
            'share_up': share_up[:10],
            'share_down': share_down[:10],
            'rank_up': rank_up[:10],
            'rank_down': rank_down[:10],
            'acos_worse': acos_worse[:5],
            'acos_better': acos_better[:5],
            'stats': {
                'total_compared': len(compared),
                'new_count': len(new_keywords),
                'lost_count': len(lost_keywords),
                'share_up_count': len(share_up),
                'share_down_count': len(share_down),
                'rank_up_count': len(rank_up),
                'rank_down_count': len(rank_down),
                'acos_worse_count': len(acos_worse),
                'acos_better_count': len(acos_better),
            },
        }

    return results


def main():
    parser = argparse.ArgumentParser(description='亚马逊搜索词Bailing分析V2')
    parser.add_argument('--excel', required=True, help='当前周期Excel文件路径')
    parser.add_argument('--excel-prev', help='上一周期Excel文件路径（传入则启用双周期对比）')
    parser.add_argument('--period-label', default='当前周期', help='当前周期标签')
    parser.add_argument('--prev-period-label', default='上周期', help='上一周期标签')
    parser.add_argument('--step', choices=['extract', 'structure', 'search_terms', 'all', 'compare'], default='all',
                        help='执行步骤：extract/structure/search_terms/all/compare')
    parser.add_argument('--output', help='输出JSON文件路径（默认stdout）')

    args = parser.parse_args()

    if args.step in ('extract', 'structure', 'all'):
        result = extract_data(args.excel)
        if 'error' in result:
            print(json.dumps(result, ensure_ascii=False))
            sys.exit(1)

        if args.step in ('extract', 'structure'):
            output = json.dumps(result, ensure_ascii=False, indent=2, default=str)
            if args.output:
                with open(args.output, 'w', encoding='utf-8') as f:
                    f.write(output)
                print(f"Saved full response: {args.output}")
            else:
                print(output)
            return

    if args.step in ('search_terms', 'all'):
        if args.step == 'all':
            portfolio_data = result
        else:
            with open(args.excel, 'r', encoding='utf-8') as f:
                portfolio_data = json.load(f)

        pareto_result = pareto_analysis(portfolio_data)

        if args.step == 'all':
            # 合并 extract + pareto 结果
            for p_name, p_data in pareto_result.items():
                if p_name in result.get('portfolios', {}):
                    result['portfolios'][p_name].update(p_data)
            output_data = result
        else:
            output_data = pareto_result

        output = json.dumps(output_data, ensure_ascii=False, indent=2, default=str)

        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(output)
            print(f"Saved full response: {args.output}")
        else:
            print(output)

    # V2: 双周期对比
    if args.excel_prev:
        prev_result = extract_data(args.excel_prev)
        if 'error' in prev_result:
            print(json.dumps(prev_result, ensure_ascii=False), file=sys.stderr)
        else:
            prev_pareto = pareto_analysis(prev_result)
            for p_name, p_data in prev_pareto.items():
                if p_name in prev_result.get('portfolios', {}):
                    prev_result['portfolios'][p_name].update(p_data)

            comparison = compare_periods(
                output_data if args.step == 'all' else result,
                prev_result,
                args.period_label, args.prev_period_label
            )

            output_data['period_comparison'] = comparison
            output_data['has_comparison'] = True

            output = json.dumps(output_data, ensure_ascii=False, indent=2, default=str)
            if args.output:
                with open(args.output, 'w', encoding='utf-8') as f:
                    f.write(output)
                print(f"Saved full response: {args.output}")
            else:
                print(output)


if __name__ == '__main__':
    main()
