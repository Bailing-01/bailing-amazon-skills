#!/usr/bin/env python3
"""
Bailing ABA ASIN Analysis Pipeline
Input: JSON with file_path (Excel path), asin (optional), brand (optional)
Output: JSON with weekly_trend, keyword_gaps, root_analysis, quadrants, ad_structure
"""
import sys
import json
import os
import re
from collections import defaultdict

# Import linkfox_paths for session-aware path resolution
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from linkfox_paths import resolve_data_path, session_root
except ImportError:
    resolve_data_path = None
    session_root = None

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": True, "message": "Missing params JSON"}))
        sys.exit(1)

    params = json.loads(sys.argv[1])
    file_path = params.get("file_path", "")
    asin = params.get("asin", "")
    brand = params.get("brand", "")

    if not file_path or not os.path.isfile(file_path):
        print(json.dumps({"error": True, "message": f"File not found: {file_path}"}))
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print(json.dumps({"error": True, "message": "openpyxl not installed"}))
        sys.exit(1)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Parse rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        date_str = row[33].strftime("%Y-%m-%d") if row[33] else "unknown"
        rows.append({
            "query": str(row[0]).strip(),
            "search_volume": row[2] or 0,
            "impressions_total": row[3] or 0,
            "asin_impressions": row[4] or 0,
            "clicks_total": row[6] or 0,
            "asin_clicks": row[8] or 0,
            "click_price_median": row[10] or 0,
            "asin_click_price": row[11] or 0,
            "cart_total": row[15] or 0,
            "asin_cart": row[17] or 0,
            "purchases_total": row[24] or 0,
            "asin_purchases": row[26] or 0,
            "purchase_price_median": row[28] or 0,
            "asin_purchase_price": row[29] or 0,
            "date": date_str,
        })

    if not rows:
        print(json.dumps({"error": True, "message": "No data rows found in Excel"}))
        sys.exit(1)

    # === 1. WEEKLY TREND ===
    weekly_market = defaultdict(lambda: {"impressions_total": 0, "clicks_total": 0, "cart_total": 0, "purchases_total": 0, "query_count": 0})
    weekly_asin = defaultdict(lambda: {"asin_impressions": 0, "asin_clicks": 0, "asin_cart": 0, "asin_purchases": 0})
    for r in rows:
        wm = weekly_market[r["date"]]
        wm["impressions_total"] += r["impressions_total"]
        wm["clicks_total"] += r["clicks_total"]
        wm["cart_total"] += r["cart_total"]
        wm["purchases_total"] += r["purchases_total"]
        wm["query_count"] += 1
        wa = weekly_asin[r["date"]]
        wa["asin_impressions"] += r["asin_impressions"]
        wa["asin_clicks"] += r["asin_clicks"]
        wa["asin_cart"] += r["asin_cart"]
        wa["asin_purchases"] += r["asin_purchases"]

    dates = sorted(weekly_market.keys())
    weekly_trend = []
    for d in dates:
        m = weekly_market[d]
        a = weekly_asin[d]
        m_ctr = m["clicks_total"] / m["impressions_total"] * 100 if m["impressions_total"] else 0
        m_conv = m["purchases_total"] / m["clicks_total"] * 100 if m["clicks_total"] else 0
        m_cart = m["cart_total"] / m["clicks_total"] * 100 if m["clicks_total"] else 0
        m_cart_conv = m["purchases_total"] / m["cart_total"] * 100 if m["cart_total"] else 0
        a_ctr = a["asin_clicks"] / a["asin_impressions"] * 100 if a["asin_impressions"] else 0
        a_conv = a["asin_purchases"] / a["asin_clicks"] * 100 if a["asin_clicks"] else 0
        a_cart = a["asin_cart"] / a["asin_clicks"] * 100 if a["asin_clicks"] else 0
        a_cart_conv = a["asin_purchases"] / a["asin_cart"] * 100 if a["asin_cart"] else 0
        a_share = a["asin_impressions"] / m["impressions_total"] * 100 if m["impressions_total"] else 0
        c_share = a["asin_clicks"] / m["clicks_total"] * 100 if m["clicks_total"] else 0
        p_share = a["asin_purchases"] / m["purchases_total"] * 100 if m["purchases_total"] else 0
        weekly_trend.append({
            "date": d,
            "market_impressions": m["impressions_total"], "market_clicks": m["clicks_total"],
            "market_cart": m["cart_total"], "market_purchases": m["purchases_total"],
            "market_ctr": round(m_ctr, 2), "market_conv": round(m_conv, 2),
            "market_cart_rate": round(m_cart, 2), "market_cart_conv": round(m_cart_conv, 2),
            "asin_impressions": a["asin_impressions"], "asin_clicks": a["asin_clicks"],
            "asin_cart": a["asin_cart"], "asin_purchases": a["asin_purchases"],
            "asin_share": round(a_share, 2), "asin_click_share": round(c_share, 2),
            "asin_purchase_share": round(p_share, 2),
            "asin_ctr": round(a_ctr, 2), "asin_conv": round(a_conv, 2),
            "asin_cart_rate": round(a_cart, 2), "asin_cart_conv": round(a_cart_conv, 2),
            "ctr_gap": round(a_ctr - m_ctr, 2), "conv_gap": round(a_conv - m_conv, 2),
        })

    # === 2. KEYWORD GAP ANALYSIS ===
    kw_data = defaultdict(lambda: {
        "impressions_total": 0, "asin_impressions": 0, "clicks_total": 0, "asin_clicks": 0,
        "cart_total": 0, "asin_cart": 0, "purchases_total": 0, "asin_purchases": 0,
        "search_volume": 0,
    })
    for r in rows:
        d = kw_data[r["query"]]
        d["search_volume"] = max(d["search_volume"], r["search_volume"])
        d["impressions_total"] += r["impressions_total"]
        d["asin_impressions"] += r["asin_impressions"]
        d["clicks_total"] += r["clicks_total"]
        d["asin_clicks"] += r["asin_clicks"]
        d["cart_total"] += r["cart_total"]
        d["asin_cart"] += r["asin_cart"]
        d["purchases_total"] += r["purchases_total"]
        d["asin_purchases"] += r["asin_purchases"]

    keyword_gaps = []
    for q, d in kw_data.items():
        a_ctr = d["asin_clicks"] / d["asin_impressions"] * 100 if d["asin_impressions"] else 0
        m_ctr = d["clicks_total"] / d["impressions_total"] * 100 if d["impressions_total"] else 0
        a_conv = d["asin_purchases"] / d["asin_clicks"] * 100 if d["asin_clicks"] else 0
        m_conv = d["purchases_total"] / d["clicks_total"] * 100 if d["clicks_total"] else 0
        a_cart = d["asin_cart"] / d["asin_clicks"] * 100 if d["asin_clicks"] else 0
        m_cart = d["cart_total"] / d["clicks_total"] * 100 if d["clicks_total"] else 0
        impr_share = d["asin_impressions"] / d["impressions_total"] * 100 if d["impressions_total"] else 0
        keyword_gaps.append({
            "query": q, "search_volume": d["search_volume"],
            "asin_impressions": d["asin_impressions"], "total_impressions": d["impressions_total"],
            "impr_share": round(impr_share, 2),
            "asin_clicks": d["asin_clicks"], "total_clicks": d["clicks_total"],
            "asin_ctr": round(a_ctr, 2), "market_ctr": round(m_ctr, 2), "ctr_gap": round(a_ctr - m_ctr, 2),
            "asin_cart": d["asin_cart"], "asin_cart_rate": round(a_cart, 2), "market_cart_rate": round(m_cart, 2),
            "asin_purchases": d["asin_purchases"], "total_purchases": d["purchases_total"],
            "asin_conv": round(a_conv, 2), "market_conv": round(m_conv, 2), "conv_gap": round(a_conv - m_conv, 2),
        })
    keyword_gaps.sort(key=lambda x: x["asin_impressions"], reverse=True)

    # === 3. ROOT WORD ANALYSIS ===
    # Dynamic brand detection
    brand_lower = brand.lower() if brand else ""
    own_brand_terms = [brand_lower] if brand_lower else []
    # Auto-detect brand from top keywords (most frequent brand-like token)
    if not own_brand_terms:
        token_freq = defaultdict(int)
        for r in keyword_gaps[:20]:
            for t in re.findall(r'[a-z]+', r["query"].lower()):
                if t not in ('neck', 'cream', 'firming', 'tighten', 'tightening', 'lift', 'lifting',
                             'serum', 'for', 'women', 'and', 'the', 'best', 'skin', 'anti', 'aging'):
                    token_freq[t] += r["asin_impressions"]
        if token_freq:
            top_brand = max(token_freq, key=token_freq.get)
            if token_freq[top_brand] > 5000:
                own_brand_terms = [top_brand]

    def tokenize(q):
        q = re.sub(r'[^a-z0-9\s]', '', q.lower().strip())
        return q.split()

    # Build root mapping
    ROOT_MAP = {
        "neck": ["neck", "nevk", "nexk", "neckcream"],
        "cream": ["cream", "creams", "crema", "cremas"],
        "firming": ["firming", "firmer", "firm"],
        "tightening": ["tightening", "tighten", "tight", "tightnes"],
        "lift": ["lift", "lifting", "lifts"],
        "serum": ["serum", "suero"],
        "wrinkle": ["wrinkle", "wrinkles"],
        "skin": ["skin", "skincare"],
        "face": ["face", "facial"],
        "women": ["women", "womens"],
        "sagging": ["sagging", "sag"],
        "chin": ["chin", "double"],
        "chest": ["chest", "declote", "decollete"],
        "turkey": ["turkey"],
        "best": ["best", "highest", "top", "rated"],
        "aging": ["aging", "age", "ageing", "anti-aging", "antiaging"],
        "korean": ["korean"],
        "hydrating": ["hydrating", "hydration", "hydrate"],
        "moisturizer": ["moisturizer", "moisturizers"],
        "amazon": ["amazon", "shopping", "online", "official", "site"],
        "ultrasound": ["ultrasound"],
        "device": ["device", "wand"],
        "stick": ["stick"],
        "gold": ["gold", "bond"],
        "roc": ["roc"],
        "monet": ["monet"],
        "strivectin": ["strivectin"],
        "olay": ["olay"],
        "murad": ["murad"],
        "crepe": ["crepe"],
        "spanish_cuello": ["cuello"],
        "spanish_papada": ["papada"],
        "spanish_crema": ["crema"],
        "spanish_para": ["para"],
        "spanish_reafirmante": ["reafirmante"],
        "spanish_grasa": ["grasa", "quemador"],
        "spanish_tensora": ["tensora", "tensor"],
        "spanish_antiarrugas": ["antiarrugas"],
        "spanish_mujer": ["mujer"],
        "spanish_manchas": ["manchas", "melasma"],
        "spanish_escote": ["escote"],
    }
    if own_brand_terms:
        ROOT_MAP["own_brand"] = own_brand_terms[:]

    token_to_root = {}
    for root_cat, variants in ROOT_MAP.items():
        for v in variants:
            token_to_root[v] = root_cat

    root_stats = defaultdict(lambda: {
        "query_count": 0, "asin_impressions": 0, "asin_clicks": 0, "asin_cart": 0, "asin_purchases": 0,
        "total_impressions": 0, "total_clicks": 0, "total_purchases": 0, "search_volume": 0,
    })
    for r in keyword_gaps:
        tokens = tokenize(r["query"])
        found_roots = set()
        for t in tokens:
            if t in token_to_root:
                found_roots.add(token_to_root[t])
        for root_cat, variants in ROOT_MAP.items():
            for v in variants:
                if v in r["query"].lower() and v not in tokens:
                    found_roots.add(root_cat)
        for root in found_roots:
            rs = root_stats[root]
            rs["query_count"] += 1
            rs["asin_impressions"] += r["asin_impressions"]
            rs["asin_clicks"] += r["asin_clicks"]
            rs["asin_cart"] += r["asin_cart"]
            rs["asin_purchases"] += r["asin_purchases"]
            rs["total_impressions"] += r["total_impressions"]
            rs["total_clicks"] += r["total_clicks"]
            rs["total_purchases"] += r["total_purchases"]
            rs["search_volume"] += r["search_volume"]

    root_analysis = []
    for root, rs in root_stats.items():
        a_ctr = rs["asin_clicks"] / rs["asin_impressions"] * 100 if rs["asin_impressions"] else 0
        m_ctr = rs["total_clicks"] / rs["total_impressions"] * 100 if rs["total_impressions"] else 0
        a_conv = rs["asin_purchases"] / rs["asin_clicks"] * 100 if rs["asin_clicks"] else 0
        m_conv = rs["total_purchases"] / rs["total_clicks"] * 100 if rs["total_clicks"] else 0
        impr_share = rs["asin_impressions"] / rs["total_impressions"] * 100 if rs["total_impressions"] else 0
        root_analysis.append({
            "root": root, "query_count": rs["query_count"], "search_volume": rs["search_volume"],
            "asin_impressions": rs["asin_impressions"], "total_impressions": rs["total_impressions"],
            "impr_share": round(impr_share, 2), "asin_clicks": rs["asin_clicks"],
            "asin_ctr": round(a_ctr, 2), "market_ctr": round(m_ctr, 2), "ctr_gap": round(a_ctr - m_ctr, 2),
            "asin_cart": rs["asin_cart"],
            "asin_purchases": rs["asin_purchases"],
            "asin_conv": round(a_conv, 2), "market_conv": round(m_conv, 2), "conv_gap": round(a_conv - m_conv, 2),
        })
    root_analysis.sort(key=lambda x: x["asin_impressions"], reverse=True)

    # === 4. QUADRANT CLASSIFICATION ===
    competitor_names = ["roc", "strivectin", "murad", "olay", "gold bond", "crepe erase",
                        "miami md", "peter thomas roth", "musely", "fabyou", "genucel",
                        "hydro nova", "idrotherapy", "crogooe", "cocoskin", "avrora",
                        "monet", "naked and thriving", "xhekpon", "new derm", "copure",
                        "virtuo", "ouhoe", "hoygi", "gkv", "dbeautypro", "aobenz",
                        "clearly basics", "bella skin", "rapid lift", "spa doctor",
                        "lebel", "praibeauty", "antivp", "doubao", "7aoo", "blanbutton",
                        "blue q", "boot tight", "beyond glow", "bio rose", "dr denise",
                        "dr lazuk", "perfect cleavage", "nevk", "nexk", "gopite", "gopitr",
                        "gopore", "gopro", "goupure", "ho pure", "get dreamy",
                        "gluta white", "thicken", "tightnes", "gold pure"]
    irrelevant_terms = ["amazon shopping", "doubao", "ultrasound", "device", "womens neck skin tightener",
                        "bella skin care wand", "eye cream", "face moisturizer", "moisturizer face cream",
                        "bio rose oil", "blue q hand cream", "grema de gobernadora", "lebel cosmetics",
                        "gluta white lotion", "suero viral", "blanbutton protector", "genucel jawline",
                        "hydro nova", "idrotherapy", "peter thomas roth", "thick cream chin",
                        "tightnes cremas", "praibeauty", "dr lazuk", "naked and thriving",
                        "firming moisturizer", "firming face cream", "amazon shopping online"]

    relevant_terms = ["neck", "cream", "firming", "tighten", "lift", "serum", "wrinkle",
                      "sagging", "chin", "throat", "chest", "declote", "contour",
                      "crema", "cuello", "papada", "reafirmante", "antiarrugas",
                      "flacido", "tensora", "turkey", "crepe", "hydrating",
                      "overnight", "korean", "anti-aging"]

    for r in keyword_gaps:
        q_lower = r["query"].lower()
        is_own = any(t in q_lower for t in own_brand_terms) if own_brand_terms else False
        is_comp = any(t in q_lower for t in competitor_names) and not is_own
        is_irr = any(t in q_lower for t in irrelevant_terms) and not is_own
        is_rel = any(t in q_lower for t in relevant_terms) or is_own

        if is_irr or (not is_rel and not is_comp and not is_own):
            r["quadrant"] = "Q4"
            r["quadrant_label"] = "不相关-否定"
        elif is_comp:
            if r["asin_purchases"] > 0:
                r["quadrant"] = "Q1"
                r["quadrant_label"] = "竞品词-表现好"
            elif r["asin_clicks"] > 5 and r["asin_conv"] < 3:
                r["quadrant"] = "Q2"
                r["quadrant_label"] = "竞品词-表现差"
            else:
                r["quadrant"] = "Q3"
                r["quadrant_label"] = "竞品词-待观察"
        elif is_own:
            if r["asin_conv"] >= 8:
                r["quadrant"] = "Q1"
                r["quadrant_label"] = "品牌词-表现好"
            else:
                r["quadrant"] = "Q2"
                r["quadrant_label"] = "品牌词-待优化"
        elif is_rel:
            if r["asin_impressions"] < 10:
                r["quadrant"] = "Q3"
                r["quadrant_label"] = "相关词-待拓展"
            elif r["asin_conv"] >= 10 and r["asin_clicks"] >= 5:
                r["quadrant"] = "Q1"
                r["quadrant_label"] = "相关词-表现好"
            elif r["asin_clicks"] >= 5 and r["asin_conv"] < 5:
                r["quadrant"] = "Q2"
                r["quadrant_label"] = "相关词-表现差"
            elif r["asin_clicks"] < 5:
                r["quadrant"] = "Q3"
                r["quadrant_label"] = "相关词-待拓展"
            else:
                r["quadrant"] = "Q1"
                r["quadrant_label"] = "相关词-表现好"
        else:
            r["quadrant"] = "Q4"
            r["quadrant_label"] = "不相关-否定"

    # Quadrant summary
    quad_summary = {}
    for qid in ["Q1", "Q2", "Q3", "Q4"]:
        items = [r for r in keyword_gaps if r["quadrant"] == qid]
        impr = sum(r["asin_impressions"] for r in items)
        clicks = sum(r["asin_clicks"] for r in items)
        purch = sum(r["asin_purchases"] for r in items)
        ctr = clicks / impr * 100 if impr else 0
        conv = purch / clicks * 100 if clicks else 0
        quad_summary[qid] = {"count": len(items), "impressions": impr, "clicks": clicks,
                             "purchases": purch, "ctr": round(ctr, 2), "conv": round(conv, 2)}

    # === 5. AD STRUCTURE ===
    q1_brand = [r for r in keyword_gaps if r["quadrant"] == "Q1" and own_brand_terms and any(t in r["query"].lower() for t in own_brand_terms)]
    q1_generic = [r for r in keyword_gaps if r["quadrant"] == "Q1" and not (own_brand_terms and any(t in r["query"].lower() for t in own_brand_terms))]
    q2_items = [r for r in keyword_gaps if r["quadrant"] == "Q2"]
    q3_items = [r for r in keyword_gaps if r["quadrant"] == "Q3"]
    q4_items = [r for r in keyword_gaps if r["quadrant"] == "Q4"]

    # Negation list
    exact_negate = [r["query"] for r in q4_items if r["asin_purchases"] == 0 and r["asin_clicks"] > 3]
    phrase_negate_candidates = [r["query"] for r in q4_items if r["asin_purchases"] <= 1]

    ad_structure = [
        {"name": "品牌词防御", "keywords": len(q1_brand), "match": "精准+词组", "bid": "高",
         "budget_pct": 30, "source": "Q1品牌词", "strategy": "维持头部排名，防御竞品抢词"},
        {"name": "核心泛词攻坚", "keywords": len(q1_generic) + len(q2_items), "match": "广泛+否定词", "bid": "中",
         "budget_pct": 20, "source": "Q1+Q2泛品类词", "strategy": "优化Listing转化后再提竞价"},
        {"name": "高转化场景词", "keywords": len([r for r in q3_items if any(t in r["query"].lower() for t in ["turkey", "over 50", "over 60", "best", "overnight", "korean"])]),
         "match": "精准", "bid": "中高", "budget_pct": 15, "source": "Q3场景词", "strategy": "低CPC高转化ROI最大化"},
        {"name": "西语市场拓展", "keywords": len([r for r in q3_items if any(t in r["query"].lower() for t in ["crema", "cuello", "papada", "para", "reafirmante"])]),
         "match": "精准+词组", "bid": "低", "budget_pct": 10, "source": "Q3西语词", "strategy": "低竞争高转化差异化获客"},
        {"name": "竞品词渗透", "keywords": len([r for r in keyword_gaps if r["quadrant"] in ("Q1", "Q3") and any(t in r["query"].lower() for t in competitor_names)]),
         "match": "精准", "bid": "低", "budget_pct": 10, "source": "Q1+Q3竞品词", "strategy": "Monet等有效竞品持续投放"},
        {"name": "自动广告探索", "keywords": 0, "match": "自动", "bid": "低",
         "budget_pct": 15, "source": "自动匹配", "strategy": "发现新词移入手动活动"},
    ]

    # Overall summary
    total_impr = sum(r["asin_impressions"] for r in keyword_gaps)
    total_clicks = sum(r["asin_clicks"] for r in keyword_gaps)
    total_purch = sum(r["asin_purchases"] for r in keyword_gaps)
    total_cart = sum(r["asin_cart"] for r in keyword_gaps)

    # Resolve data path for intermediate data persistence
    data_dir = None
    if resolve_data_path:
        try:
            data_path = resolve_data_path("bailing-aba-asin-analysis", str(int(__import__("time").time())))
            data_dir = os.path.dirname(data_path)
            os.makedirs(data_dir, exist_ok=True)
        except Exception:
            pass

    result = {
        "error": False,
        "asin": asin,
        "brand": brand or (own_brand_terms[0] if own_brand_terms else ""),
        "date_range": f"{dates[0]} ~ {dates[-1]}" if dates else "",
        "data_dir": data_dir,
        "total_queries": len(keyword_gaps),
        "total_impressions": total_impr,
        "total_clicks": total_clicks,
        "total_cart": total_cart,
        "total_purchases": total_purch,
        "avg_ctr": round(total_clicks / total_impr * 100, 2) if total_impr else 0,
        "avg_conv": round(total_purch / total_clicks * 100, 2) if total_clicks else 0,
        "weekly_trend": weekly_trend,
        "keyword_gaps": keyword_gaps,
        "root_analysis": root_analysis,
        "quadrant_summary": quad_summary,
        "ad_structure": ad_structure,
        "negation": {"exact": exact_negate, "phrase_candidates": phrase_negate_candidates},
        "q1_keywords": [r for r in keyword_gaps if r["quadrant"] == "Q1"][:30],
        "q2_keywords": [r for r in keyword_gaps if r["quadrant"] == "Q2"][:20],
        "q3_keywords": sorted([r for r in keyword_gaps if r["quadrant"] == "Q3"], key=lambda x: x["asin_purchases"], reverse=True)[:30],
        "q4_keywords": sorted([r for r in keyword_gaps if r["quadrant"] == "Q4"], key=lambda x: x["asin_clicks"], reverse=True)[:30],
    }

    print(json.dumps(result, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
